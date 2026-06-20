from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidgetItem,
                               QPushButton, QLabel, QGroupBox, QSplitter,
                               QComboBox, QDateEdit, QLineEdit, QCheckBox,
                               QTextEdit, QListWidget, QListWidgetItem,
                               QDoubleSpinBox, QSpinBox, QDialog, QInputDialog,
                               QTabWidget, QFormLayout, QHeaderView)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor, QBrush, QFont
from ui.base_widget import BaseWidget, BaseDialog
from modules.bill_service import BillService
from modules.discount_service import DiscountService
from modules.member_service import MemberService
from database.models import CouponType, MemberLevel
from datetime import datetime, date, timedelta
from datetime import date as py_date


class BillingDialog(BaseDialog):
    def __init__(self, parent=None, db=None, booking=None):
        self.db = db
        self.booking = booking
        self.selected_coupons = []
        self.current_result = None
        self.created_bill = None
        self.selected_member_id = None
        super().__init__("账单结算", parent)
        self.discount_service = DiscountService(db)
        self.bill_service = BillService(db)
        self.member_service = MemberService(db)
        self.resize(750, 700)
        self._load_data()

    def init_dialog(self):
        main_layout = QVBoxLayout()

        info_group = QGroupBox("预订信息")
        info_layout = QFormLayout(info_group)
        self.lbl_table = QLabel("")
        self.lbl_customer = QLabel("")
        self.lbl_date = QLabel("")
        self.lbl_time = QLabel("")
        self.lbl_hours = QLabel("")
        self.lbl_base = QLabel("")
        self.lbl_base.setStyleSheet("font-size: 16px; font-weight: bold;")
        info_layout.addRow("桌号:", self.lbl_table)
        info_layout.addRow("客人:", self.lbl_customer)
        info_layout.addRow("日期:", self.lbl_date)
        info_layout.addRow("时间:", self.lbl_time)
        info_layout.addRow("时长:", self.lbl_hours)
        info_layout.addRow("基础金额:", self.lbl_base)
        main_layout.addWidget(info_group)

        coupon_group = QGroupBox("选择优惠券")
        coupon_layout = QVBoxLayout(coupon_group)

        list_splitter = QSplitter(Qt.Horizontal)

        available_group = QGroupBox("可用优惠券")
        available_layout = QVBoxLayout(available_group)
        self.list_available = QListWidget()
        self.list_available.setSelectionMode(QListWidget.MultiSelection)
        available_layout.addWidget(self.list_available)
        list_splitter.addWidget(available_group)

        selected_group = QGroupBox("已选优惠券")
        selected_layout = QVBoxLayout(selected_group)
        self.list_selected = QListWidget()
        self.list_selected.setSelectionMode(QListWidget.MultiSelection)
        selected_layout.addWidget(self.list_selected)
        list_splitter.addWidget(selected_group)

        coupon_layout.addWidget(list_splitter)

        btn_layout = QHBoxLayout()
        self.btn_add_coupon = QPushButton("→ 添加")
        self.btn_remove_coupon = QPushButton("← 移除")
        self.btn_clear_coupon = QPushButton("清空")
        self.btn_calc = QPushButton("🧮 计算")
        self.btn_add_coupon.clicked.connect(self.add_coupon)
        self.btn_remove_coupon.clicked.connect(self.remove_coupon)
        self.btn_clear_coupon.clicked.connect(self.clear_coupons)
        self.btn_calc.clicked.connect(self.calculate)
        btn_layout.addWidget(self.btn_add_coupon)
        btn_layout.addWidget(self.btn_remove_coupon)
        btn_layout.addWidget(self.btn_clear_coupon)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_calc)
        coupon_layout.addLayout(btn_layout)

        main_layout.addWidget(coupon_group)

        result_group = QGroupBox("计算结果")
        result_layout = QVBoxLayout(result_group)
        self.result_text = QTextEdit()
        self.result_text.setReadOnly(True)
        self.result_text.setStyleSheet("font-family: Consolas, Monaco, monospace; font-size: 13px;")
        self.result_text.setMaximumHeight(120)
        result_layout.addWidget(self.result_text)

        amount_layout = QHBoxLayout()
        amount_layout.addWidget(QLabel("优惠合计:"))
        self.lbl_discount = QLabel("¥0.00")
        self.lbl_discount.setStyleSheet("font-size: 14px; color: #f44336; font-weight: bold;")
        amount_layout.addWidget(self.lbl_discount)
        amount_layout.addStretch()
        amount_layout.addWidget(QLabel("应付金额:"))
        self.lbl_final = QLabel("¥0.00")
        self.lbl_final.setStyleSheet("font-size: 18px; color: #1976d2; font-weight: bold;")
        amount_layout.addWidget(self.lbl_final)
        result_layout.addLayout(amount_layout)

        main_layout.addWidget(result_group)

        pay_group = QGroupBox("支付信息")
        pay_layout = QFormLayout(pay_group)
        pay_method_layout = QHBoxLayout()
        self.combo_payment = QComboBox()
        self.combo_payment.addItems(["现金", "微信支付", "支付宝", "银行卡", "会员卡", "其他"])
        self.combo_payment.setMinimumHeight(32)
        self.combo_payment.currentTextChanged.connect(self.on_payment_method_changed)
        pay_method_layout.addWidget(self.combo_payment)
        pay_method_layout.addStretch()
        pay_layout.addRow("支付方式:", pay_method_layout)

        member_layout = QHBoxLayout()
        self.combo_member = QComboBox()
        self.combo_member.setMinimumHeight(32)
        member_layout.addWidget(self.combo_member, 1)
        self.lbl_member_balance = QLabel("")
        self.lbl_member_balance.setStyleSheet("color: #1976d2; font-weight: bold;")
        member_layout.addWidget(self.lbl_member_balance)
        pay_layout.addRow("选择会员:", member_layout)

        self.chk_print = QCheckBox("打印账单")
        self.chk_print.setChecked(True)
        pay_layout.addRow("", self.chk_print)

        main_layout.addWidget(pay_group)

        self.form_layout.addRow(main_layout)

    def _load_data(self):
        if not self.booking:
            return

        try:
            table = self.booking.table
            if table:
                self.lbl_table.setText(f"{table.table_number} - {table.name}")
            self.lbl_customer.setText(self.booking.customer_name or "")
            self.lbl_date.setText(str(self.booking.booking_date))
            time_str = f"{self.booking.start_time.strftime('%H:%M')}-{self.booking.end_time.strftime('%H:%M')}"
            self.lbl_time.setText(time_str)
            self.lbl_hours.setText(f"{self.booking.total_hours:.1f} 小时")
            self.lbl_base.setText(f"¥{self.booking.base_amount:.2f}")

            coupons = self.discount_service.get_available_coupons(self.booking.base_amount)
            for coupon in coupons:
                if coupon.type == CouponType.DISCOUNT:
                    desc = f"{coupon.name} - {coupon.discount_value}折 (满{coupon.min_consumption or 0}元)"
                else:
                    desc = f"{coupon.name} - 减{coupon.discount_value}元 (满{coupon.min_consumption or 0}元)"
                item = QListWidgetItem(desc)
                item.setData(Qt.UserRole, coupon.id)
                self.list_available.addItem(item)

            members = self.member_service.get_all_members(only_active=True)
            self.combo_member.addItem("不使用会员（无会员折扣）", None)
            for m in members:
                level_discount = self.member_service.get_level_discount(m.level)
                level_text = m.level.value if hasattr(m.level, 'value') else str(m.level)
                desc = f"{m.name} ({m.phone}) - {level_text}({level_discount}折) 余额:¥{m.balance:.2f}"
                self.combo_member.addItem(desc, m.id)

            self.combo_member.currentIndexChanged.connect(self.on_member_changed)
            self.calculate()
        except Exception as e:
            self.result_text.setPlainText(f"加载出错: {str(e)}")

    def on_payment_method_changed(self, method):
        is_member = method == "会员卡"
        if is_member:
            self._update_member_balance()
        else:
            self.lbl_member_balance.setText("")

    def on_member_changed(self):
        self._update_member_balance()
        self.calculate()

    def _update_member_balance(self):
        idx = self.combo_member.currentIndex()
        if idx >= 0:
            member_id = self.combo_member.itemData(idx)
            if member_id:
                member = self.member_service.get_member(member_id)
                if member:
                    self.lbl_member_balance.setText(f"余额: ¥{member.balance:.2f}")
                    return
        self.lbl_member_balance.setText("")

    def add_coupon(self):
        for item in self.list_available.selectedItems():
            coupon_id = item.data(Qt.UserRole)
            if coupon_id not in self.selected_coupons:
                self.selected_coupons.append(coupon_id)
                new_item = QListWidgetItem(item.text())
                new_item.setData(Qt.UserRole, coupon_id)
                self.list_selected.addItem(new_item)
        self.calculate()

    def remove_coupon(self):
        for item in self.list_selected.selectedItems():
            coupon_id = item.data(Qt.UserRole)
            if coupon_id in self.selected_coupons:
                self.selected_coupons.remove(coupon_id)
            row = self.list_selected.row(item)
            self.list_selected.takeItem(row)
        self.calculate()

    def clear_coupons(self):
        self.selected_coupons.clear()
        self.list_selected.clear()
        self.calculate()

    def _get_selected_member_id(self):
        idx = self.combo_member.currentIndex()
        if idx >= 0:
            return self.combo_member.itemData(idx)
        return None

    def calculate(self):
        if not self.booking:
            return
        try:
            member_id = self._get_selected_member_id()
            result = self.discount_service.calculate_discount(
                self.booking.base_amount,
                coupon_ids=self.selected_coupons,
                member_id=member_id
            )

            lines = list(result.calculation_steps) if result.calculation_steps else []
            lines.append("")
            lines.append("=" * 50)
            lines.append(f"基础金额: ¥{result.base_amount:.2f}")
            if hasattr(result, 'member_discount') and result.member_discount > 0:
                lines.append(f"会员折扣: -¥{result.member_discount:.2f}")
            if hasattr(result, 'coupon_discount') and result.coupon_discount > 0:
                lines.append(f"优惠券折扣: -¥{result.coupon_discount:.2f}")
            lines.append(f"优惠合计: ¥{result.discount_amount:.2f}")
            lines.append(f"最终应付: ¥{result.final_amount:.2f}")
            if result.has_negative_protection:
                lines.append("⚠️ 已触发负值兜底校验，金额已调整为0")

            if result.applied_discounts:
                lines.append("")
                lines.append("优惠明细:")
                for d in result.applied_discounts:
                    lines.append(f"  [{d.get('apply_order', '?')}] {d.get('coupon_name', '')} ({d.get('coupon_type', '')}): -¥{d.get('applied_amount', 0):.2f}")

            self.result_text.setPlainText("\n".join(lines))
            self.lbl_discount.setText(f"-¥{result.discount_amount:.2f}")
            self.lbl_final.setText(f"¥{result.final_amount:.2f}")
            self.current_result = result
        except Exception as e:
            self.result_text.setPlainText(f"计算出错: {str(e)}")
            self.current_result = None

    def validate(self):
        if not self.booking:
            self.show_error("错误", "预订信息不存在")
            return False
        if not self.current_result:
            self.show_error("错误", "请先计算优惠金额")
            return False

        if self.combo_payment.currentText() == "会员卡":
            member_id = self._get_selected_member_id()
            if not member_id:
                self.show_error("错误", "请选择会员")
                return False
            member = self.member_service.get_member(member_id)
            if not member:
                self.show_error("错误", "会员不存在")
                return False
            if member.balance < self.current_result.final_amount:
                self.show_error("余额不足",
                    f"会员「{member.name}」余额 ¥{member.balance:.2f}，"
                    f"应付 ¥{self.current_result.final_amount:.2f}，"
                    f"差额 ¥{self.current_result.final_amount - member.balance:.2f}\n"
                    f"请先充值或更换支付方式！")
                return False

        return True

    def accept(self):
        if not self.validate():
            return

        try:
            bill = self.bill_service.create_bill(
                booking_id=self.booking.id,
                discount_result=self.current_result
            )

            payment_method = self.combo_payment.currentText()
            member_id = None
            if payment_method == "会员卡":
                idx = self.combo_member.currentIndex()
                member_id = self.combo_member.itemData(idx)

            self.bill_service.pay_bill(bill.id, payment_method, member_id=member_id)

            self.created_bill = bill

            if self.chk_print.isChecked():
                try:
                    print_content = self.bill_service.generate_print_content(bill.id)
                    from PySide6.QtPrintSupport import QPrinter, QPrintDialog
                    from PySide6.QtGui import QTextDocument

                    printer = QPrinter(QPrinter.HighResolution)
                    printer.setPageSize(QPrinter.A4)
                    printer.setOutputFormat(QPrinter.NativeFormat)
                    dialog = QPrintDialog(printer, self)
                    if dialog.exec() == QDialog.Accepted:
                        doc = QTextDocument()
                        doc.setPlainText(print_content)
                        doc.print_(printer)
                except ImportError:
                    pass
                except Exception:
                    pass

            self.show_info("成功", f"账单已生成并支付成功！\n账单号: {bill.bill_no}\n应付: ¥{bill.final_amount:.2f}")
            super().accept()
        except Exception as e:
            self.show_error("错误", str(e))


class BillingWidget(BaseWidget):
    def __init__(self, db, parent=None):
        super().__init__(db, parent)
        self.bill_service = BillService(self.db)
        self.discount_service = DiscountService(self.db)
        self.refresh()

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)

        title = QLabel("� 账单与营业管理")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #1976d2;")
        main_layout.addWidget(title)

        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabBar::tab {
                padding: 8px 20px;
                font-size: 13px;
            }
            QTabBar::tab:selected {
                background: white;
                font-weight: bold;
                color: #1976d2;
            }
        """)

        self.tabs.addTab(self._create_bill_tab(), "📋 账单列表")
        self.tabs.addTab(self._create_stats_tab(), "📊 营业统计")

        main_layout.addWidget(self.tabs)

    def _create_bill_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(8)

        filter_group = QGroupBox("查询条件")
        filter_layout = QHBoxLayout(filter_group)

        filter_layout.addWidget(QLabel("开始日期:"))
        self.date_start = QDateEdit()
        self.date_start.setCalendarPopup(True)
        self.date_start.setDisplayFormat("yyyy-MM-dd")
        self.date_start.setDate(QDate.currentDate().addDays(-7))
        self.date_start.dateChanged.connect(self.refresh)
        filter_layout.addWidget(self.date_start)

        filter_layout.addWidget(QLabel("结束日期:"))
        self.date_end = QDateEdit()
        self.date_end.setCalendarPopup(True)
        self.date_end.setDisplayFormat("yyyy-MM-dd")
        self.date_end.setDate(QDate.currentDate())
        self.date_end.dateChanged.connect(self.refresh)
        filter_layout.addWidget(self.date_end)

        self.chk_only_unpaid = QCheckBox("仅未支付")
        self.chk_only_unpaid.toggled.connect(self.refresh)
        filter_layout.addWidget(self.chk_only_unpaid)

        self.btn_refresh = self.create_button("🔄 刷新", callback=self.refresh)
        filter_layout.addWidget(self.btn_refresh)
        filter_layout.addStretch()

        self.lbl_stats = QLabel("")
        self.lbl_stats.setStyleSheet("font-weight: bold; color: #666;")
        filter_layout.addWidget(self.lbl_stats)

        layout.addWidget(filter_group)

        bill_group = QGroupBox("账单列表")
        bill_layout = QVBoxLayout(bill_group)

        btn_layout = QHBoxLayout()
        self.btn_view = self.create_button("👁️ 详情", callback=self.view_bill)
        self.btn_print = self.create_button("🖨️ 打印", callback=self.print_bill)
        self.btn_pay = self.create_button("💰 支付", callback=self.pay_bill)
        self.btn_delete = self.create_button("🗑️ 删除", callback=self.delete_bill)
        btn_layout.addWidget(self.btn_view)
        btn_layout.addWidget(self.btn_print)
        btn_layout.addWidget(self.btn_pay)
        btn_layout.addWidget(self.btn_delete)
        btn_layout.addStretch()
        bill_layout.addLayout(btn_layout)

        self.bill_widget = self.create_table([
            "ID", "账单号", "桌号", "客人", "日期", "基础金额", "优惠金额", "应付金额", "支付方式", "状态"
        ])
        bill_layout.addWidget(self.bill_widget)

        layout.addWidget(bill_group)

        self.bill_widget.itemSelectionChanged.connect(self.on_bill_selected)

        return tab

    def _create_stats_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(8)

        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("视图模式:"))
        self.combo_view_mode = QComboBox()
        self.combo_view_mode.addItems(["日视图", "周视图", "月视图"])
        self.combo_view_mode.setMinimumHeight(32)
        self.combo_view_mode.currentTextChanged.connect(self.on_view_mode_changed)
        filter_layout.addWidget(self.combo_view_mode)

        filter_layout.addSpacing(20)
        filter_layout.addWidget(QLabel("统计周期:"))
        self.combo_period = QComboBox()
        self.combo_period.addItems(["近7天", "近30天", "本月", "上月", "自定义"])
        self.combo_period.currentTextChanged.connect(self.on_period_changed)
        self.combo_period.setMinimumHeight(32)
        filter_layout.addWidget(self.combo_period)

        filter_layout.addWidget(QLabel("自定义:"))
        self.stats_date_start = QDateEdit()
        self.stats_date_start.setCalendarPopup(True)
        self.stats_date_start.setDisplayFormat("yyyy-MM-dd")
        self.stats_date_start.setDate(QDate.currentDate().addDays(-7))
        filter_layout.addWidget(self.stats_date_start)
        filter_layout.addWidget(QLabel("至"))
        self.stats_date_end = QDateEdit()
        self.stats_date_end.setCalendarPopup(True)
        self.stats_date_end.setDisplayFormat("yyyy-MM-dd")
        self.stats_date_end.setDate(QDate.currentDate())
        filter_layout.addWidget(self.stats_date_end)

        self.btn_stats = self.create_button("📊 生成统计", callback=self.generate_statistics)
        filter_layout.addWidget(self.btn_stats)
        self.btn_export = self.create_button("📥 导出Excel", callback=self.export_excel)
        filter_layout.addWidget(self.btn_export)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        summary_group = QGroupBox("汇总数据")
        summary_layout = QHBoxLayout(summary_group)
        self.lbl_stat_bills = QLabel("账单数: 0")
        self.lbl_stat_base = QLabel("营业额: ¥0")
        self.lbl_stat_discount = QLabel("优惠额: ¥0")
        self.lbl_stat_final = QLabel("实收额: ¥0")
        self.lbl_stat_hours = QLabel("使用时长: 0h")
        self.lbl_stat_avg = QLabel("客单价: ¥0")
        for lbl in [self.lbl_stat_bills, self.lbl_stat_base, self.lbl_stat_discount,
                     self.lbl_stat_final, self.lbl_stat_hours, self.lbl_stat_avg]:
            lbl.setStyleSheet("font-size: 14px; font-weight: bold; padding: 8px; background: white; border-radius: 4px;")
            summary_layout.addWidget(lbl)
        layout.addWidget(summary_group)

        self.stats_tabs = QTabWidget()
        self.stats_tabs.addTab(self._create_report_view(), "📋 报表视图")
        self.stats_tabs.addTab(self._create_dashboard_view(), "📈 老板看板")
        self.stats_tabs.currentChanged.connect(self.on_stats_tab_changed)
        layout.addWidget(self.stats_tabs)

        self.current_stats_data = None
        self.current_analysis_data = None
        return tab

    def _create_report_view(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        stats_splitter = QSplitter(Qt.Horizontal)

        daily_group = QGroupBox("周期明细")
        daily_layout = QVBoxLayout(daily_group)
        self.period_table = self.create_table([
            "周期", "账单数", "营业额", "优惠额", "实收额", "使用时长", "客单价"
        ])
        daily_layout.addWidget(self.period_table)
        stats_splitter.addWidget(daily_group)

        payment_group = QGroupBox("支付方式占比")
        payment_layout = QVBoxLayout(payment_group)
        self.payment_table = self.create_table([
            "支付方式", "笔数", "金额", "时长", "占比"
        ])
        payment_layout.addWidget(self.payment_table)
        stats_splitter.addWidget(payment_group)

        stats_splitter.setSizes([500, 300])
        layout.addWidget(stats_splitter)
        return tab

    def _create_dashboard_view(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(8)

        kpi_group = QGroupBox("经营概览")
        kpi_grid = QHBoxLayout(kpi_group)
        self.dash_utilization = self._create_kpi_label("桌台利用率", "0%")
        self.dash_member_ratio = self._create_kpi_label("会员消费占比", "0%")
        self.dash_member_amount_ratio = self._create_kpi_label("会员金额占比", "0%")
        self.dash_tables = self._create_kpi_label("桌台总数", "0张")
        self.dash_days = self._create_kpi_label("统计天数", "0天")
        self.dash_avg = self._create_kpi_label("客单价", "¥0")
        for kpi in [self.dash_utilization, self.dash_member_ratio, self.dash_member_amount_ratio,
                    self.dash_tables, self.dash_days, self.dash_avg]:
            kpi_grid.addWidget(kpi)
        layout.addWidget(kpi_group)

        dash_splitter = QSplitter(Qt.Horizontal)

        table_util_group = QGroupBox("桌台利用率明细")
        table_util_layout = QVBoxLayout(table_util_group)
        self.table_util_table = self.create_table([
            "桌号", "消费次数", "营收", "使用时长", "利用率估算"
        ])
        table_util_layout.addWidget(self.table_util_table)
        dash_splitter.addWidget(table_util_group)

        right_col = QWidget()
        right_layout = QVBoxLayout(right_col)

        hour_group = QGroupBox("热门时段")
        hour_layout = QVBoxLayout(hour_group)
        self.hot_hour_table = self.create_table(["时段", "次数", "营收"])
        hour_layout.addWidget(self.hot_hour_table)
        right_layout.addWidget(hour_group)

        discount_group = QGroupBox("优惠来源占比")
        discount_layout = QVBoxLayout(discount_group)
        self.discount_source_table = self.create_table(["优惠来源", "使用笔数", "优惠金额", "占比"])
        discount_layout.addWidget(self.discount_source_table)
        right_layout.addWidget(discount_group)

        dash_splitter.addWidget(right_col)
        dash_splitter.setSizes([450, 350])
        layout.addWidget(dash_splitter)

        return tab

    def _create_kpi_label(self, title: str, value: str) -> QWidget:
        box = QWidget()
        box.setStyleSheet("background: white; border-radius: 6px; border: 1px solid #e0e0e0; padding: 4px;")
        v = QVBoxLayout(box)
        v.setContentsMargins(10, 6, 10, 6)
        lbl_title = QLabel(title)
        lbl_title.setStyleSheet("font-size: 12px; color: #757575;")
        lbl_value = QLabel(value)
        lbl_value.setStyleSheet("font-size: 20px; font-weight: bold; color: #1976d2;")
        v.addWidget(lbl_title)
        v.addWidget(lbl_value)
        return box

    def _set_kpi_value(self, kpi_widget: QWidget, value: str, color: str = None):
        lbl_value = None
        for c in kpi_widget.findChildren(QLabel):
            if c.styleSheet() and "font-size: 20px" in c.styleSheet():
                lbl_value = c
                break
        if lbl_value:
            if color:
                lbl_value.setStyleSheet(f"font-size: 20px; font-weight: bold; color: {color};")
            lbl_value.setText(value)

    def on_stats_tab_changed(self, idx):
        if self.current_analysis_data and idx == 1:
            self._render_dashboard()

    def on_view_mode_changed(self, text):
        if self.current_stats_data:
            self._update_period_display()

    def on_period_changed(self, text):
        today = QDate.currentDate()
        if text == "近7天":
            self.stats_date_start.setDate(today.addDays(-6))
            self.stats_date_end.setDate(today)
        elif text == "近30天":
            self.stats_date_start.setDate(today.addDays(-29))
            self.stats_date_end.setDate(today)
        elif text == "本月":
            self.stats_date_start.setDate(QDate(today.year(), today.month(), 1))
            self.stats_date_end.setDate(today)
        elif text == "上月":
            first_of_this = QDate(today.year(), today.month(), 1)
            self.stats_date_start.setDate(first_of_this.addMonths(-1))
            self.stats_date_end.setDate(first_of_this.addDays(-1))

    def generate_statistics(self):
        start_date = self.stats_date_start.date().toPython()
        end_date = self.stats_date_end.date().toPython()

        stats = self.bill_service.get_statistics(start_date, end_date)
        self.current_stats_data = stats

        try:
            analysis = self.bill_service.get_business_analysis(start_date, end_date)
            self.current_analysis_data = analysis
        except Exception as e:
            print(f"Warning: failed to get analysis: {e}")
            self.current_analysis_data = None

        self.lbl_stat_bills.setText(f"账单数: {stats['bill_count']}")
        self.lbl_stat_base.setText(f"营业额: ¥{stats['total_base']:.2f}")
        self.lbl_stat_discount.setText(f"优惠额: ¥{stats['total_discount']:.2f}")
        self.lbl_stat_final.setText(f"实收额: ¥{stats['total_final']:.2f}")
        self.lbl_stat_hours.setText(f"使用时长: {stats['total_hours']:.1f}h")
        self.lbl_stat_avg.setText(f"客单价: ¥{stats['avg_per_bill']:.2f}")

        self._update_period_display()

        self.clear_table(self.payment_table)
        total_amount = stats["total_final"] if stats["total_final"] > 0 else 1
        for method, data in sorted(stats["by_payment"].items()):
            pct = round(data["amount"] / total_amount * 100, 1)
            row = self.payment_table.rowCount()
            self.payment_table.insertRow(row)
            self.set_table_row(self.payment_table, row, [
                method, data["count"], f"¥{data['amount']:.2f}",
                f"{data['hours']:.1f}h", f"{pct}%"
            ])

        if self.current_analysis_data and self.stats_tabs.currentIndex() == 1:
            self._render_dashboard()

    def _render_dashboard(self):
        data = self.current_analysis_data
        if not data:
            return

        self._set_kpi_value(self.dash_utilization, f"{data['utilization']}%")
        mem_color = "#4caf50" if data['member_ratio'] >= 50 else "#ff9800"
        self._set_kpi_value(self.dash_member_ratio, f"{data['member_ratio']}%", mem_color)
        self._set_kpi_value(self.dash_member_amount_ratio, f"{data['member_amount_ratio']}%")
        self._set_kpi_value(self.dash_tables, f"{data['total_tables']}张")
        self._set_kpi_value(self.dash_days, f"{data['date_range_days']}天")
        self._set_kpi_value(self.dash_avg, f"¥{data['avg_per_bill']:.0f}")

        self.clear_table(self.table_util_table)
        by_table = data["by_table"]
        possible_per_table = data["date_range_days"] * 12
        for tn in sorted(by_table.keys()):
            t = by_table[tn]
            util_e = round(t["hours"] / possible_per_table * 100, 1) if possible_per_table > 0 else 0
            row = self.table_util_table.rowCount()
            self.table_util_table.insertRow(row)
            self.set_table_row(self.table_util_table, row, [
                tn, t["count"], f"¥{t['final']:.2f}", f"{t['hours']:.1f}h", f"{util_e}%"
            ])
            util_item = self.table_util_table.item(row, 4)
            if util_e >= 70:
                util_item.setForeground(QBrush(QColor("#4caf50")))
            elif util_e >= 40:
                util_item.setForeground(QBrush(QColor("#ff9800")))
            else:
                util_item.setForeground(QBrush(QColor("#9e9e9e")))

        self.clear_table(self.hot_hour_table)
        for i, hh in enumerate(data["hot_hours"]):
            row = self.hot_hour_table.rowCount()
            self.hot_hour_table.insertRow(row)
            self.set_table_row(self.hot_hour_table, row, [
                hh["hour_range"], hh["count"], f"¥{hh['amount']:.2f}"
            ])
            if i == 0:
                self.hot_hour_table.item(row, 0).setForeground(QBrush(QColor("#f44336")))

        self.clear_table(self.discount_source_table)
        total_discount = sum(d["amount"] for d in data["by_discount_source"].values())
        total_discount = total_discount if total_discount > 0 else 1
        for src in ["会员折扣", "优惠券折扣"]:
            d = data["by_discount_source"].get(src, {"count": 0, "amount": 0})
            pct = round(d["amount"] / total_discount * 100, 1)
            row = self.discount_source_table.rowCount()
            self.discount_source_table.insertRow(row)
            self.set_table_row(self.discount_source_table, row, [
                src, d["count"], f"¥{d['amount']:.2f}", f"{pct}%"
            ])
            if src == "会员折扣":
                self.discount_source_table.item(row, 0).setForeground(QBrush(QColor("#9c27b0")))
            else:
                self.discount_source_table.item(row, 0).setForeground(QBrush(QColor("#2196f3")))

    def _update_period_display(self):
        if not self.current_stats_data:
            return

        view_mode = self.combo_view_mode.currentText()
        by_date = self.current_stats_data["by_date"]

        if view_mode == "日视图":
            self._display_daily_view(by_date)
        elif view_mode == "周视图":
            self._display_weekly_view(by_date)
        elif view_mode == "月视图":
            self._display_monthly_view(by_date)

    def _display_daily_view(self, by_date):
        self.clear_table(self.period_table)
        for ds in sorted(by_date.keys()):
            d = by_date[ds]
            avg = round(d["final"] / d["count"], 2) if d["count"] > 0 else 0
            row = self.period_table.rowCount()
            self.period_table.insertRow(row)
            self.set_table_row(self.period_table, row, [
                ds, d["count"], f"¥{d['base']:.2f}", f"¥{d['discount']:.2f}",
                f"¥{d['final']:.2f}", f"{d['hours']:.1f}h", f"¥{avg:.2f}"
            ])

    def _display_weekly_view(self, by_date):
        self.clear_table(self.period_table)
        from datetime import timedelta as td
        weeks = {}
        for ds in sorted(by_date.keys()):
            d = date.fromisoformat(ds)
            monday = d - td(days=d.weekday())
            sunday = monday + td(days=6)
            week_key = f"{monday.isoformat()} ~ {sunday.isoformat()}"
            if week_key not in weeks:
                weeks[week_key] = {"count": 0, "base": 0.0, "discount": 0.0,
                                    "final": 0.0, "hours": 0.0}
            for k in ["count", "base", "discount", "final", "hours"]:
                weeks[week_key][k] += by_date[ds][k]

        for week_key in sorted(weeks.keys()):
            w = weeks[week_key]
            avg = round(w["final"] / w["count"], 2) if w["count"] > 0 else 0
            row = self.period_table.rowCount()
            self.period_table.insertRow(row)
            self.set_table_row(self.period_table, row, [
                week_key, w["count"], f"¥{w['base']:.2f}", f"¥{w['discount']:.2f}",
                f"¥{w['final']:.2f}", f"{w['hours']:.1f}h", f"¥{avg:.2f}"
            ])

    def _display_monthly_view(self, by_date):
        self.clear_table(self.period_table)
        months = {}
        for ds in sorted(by_date.keys()):
            month_key = ds[:7]
            if month_key not in months:
                months[month_key] = {"count": 0, "base": 0.0, "discount": 0.0,
                                      "final": 0.0, "hours": 0.0}
            for k in ["count", "base", "discount", "final", "hours"]:
                months[month_key][k] += by_date[ds][k]

        for month_key in sorted(months.keys()):
            m = months[month_key]
            avg = round(m["final"] / m["count"], 2) if m["count"] > 0 else 0
            row = self.period_table.rowCount()
            self.period_table.insertRow(row)
            self.set_table_row(self.period_table, row, [
                month_key, m["count"], f"¥{m['base']:.2f}", f"¥{m['discount']:.2f}",
                f"¥{m['final']:.2f}", f"{m['hours']:.1f}h", f"¥{avg:.2f}"
            ])

    def export_excel(self):
        if not self.current_stats_data:
            self.show_info("提示", "请先生成统计数据")
            return

        try:
            from PySide6.QtWidgets import QFileDialog
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            default_name = f"营业报表_{self.stats_date_start.date().toString('yyyyMMdd')}-{self.stats_date_end.date().toString('yyyyMMdd')}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self, "导出Excel报表", default_name, "Excel文件 (*.xlsx)"
            )
            if not file_path:
                return

            stats = self.current_stats_data
            analysis = self.current_analysis_data
            wb = Workbook()

            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            title_font = Font(bold=True, size=16)
            period_font = Font(size=11, italic=True)
            section_font = Font(bold=True, size=12)

            def _style_header_row(ws, headers, row=1):
                for i, h in enumerate(headers):
                    cell = ws.cell(row=row, column=i + 1, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

            def _style_data_rows(ws, rows_data, start_row, num_fmt_cols=None):
                num_fmt_cols = num_fmt_cols or []
                for ri, row_vals in enumerate(rows_data):
                    r = start_row + ri
                    for ci, v in enumerate(row_vals):
                        cell = ws.cell(row=r, column=ci + 1, value=v)
                        cell.border = thin_border
                        cell.alignment = center_align
                        if ci in num_fmt_cols and isinstance(v, (int, float)):
                            cell.number_format = '¥#,##0.00'

            def _write_title(ws, title, period, cols=6):
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
                ws.cell(row=1, column=1, value=title).font = title_font
                ws.cell(row=1, column=1).alignment = center_align
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
                ws.cell(row=2, column=1, value=f"统计周期: {period}").font = period_font
                ws.cell(row=2, column=1).alignment = center_align

            def _auto_width(ws, cols_num, min_w=14):
                for col in range(1, cols_num + 1):
                    ws.column_dimensions[chr(64 + col)].width = max(min_w, 18)

            period_str = f"{stats['start_date']} 至 {stats['end_date']}"

            # ============ Sheet1: 营业汇总 ============
            ws1 = wb.active
            ws1.title = "营业汇总"
            _write_title(ws1, "茶楼麻将房 - 营业汇总表", period_str, 6)

            ws1.cell(row=4, column=1, value="核心指标").font = section_font

            summary_headers = ["指标", "数值"]
            _style_header_row(ws1, summary_headers, row=5)

            summary_data = [
                ["账单总数", stats['bill_count']],
                ["营业额（应收）", round(stats['total_base'], 2)],
                ["优惠合计", round(stats['total_discount'], 2)],
                ["实收金额", round(stats['total_final'], 2)],
                ["使用时长（小时）", round(stats['total_hours'], 1)],
                ["客单价", round(stats['avg_per_bill'], 2)],
            ]
            _style_data_rows(ws1, summary_data, start_row=6, num_fmt_cols=[1])

            start_row = 6 + len(summary_data) + 2
            ws1.cell(row=start_row, column=1, value="周期明细").font = section_font
            start_row += 1

            view_mode = self.combo_view_mode.currentText()
            detail_headers = ["周期", "账单数", "营业额", "优惠额", "实收额", "使用时长(h)", "客单价"]
            _style_header_row(ws1, detail_headers, row=start_row)
            start_row += 1

            period_rows = []
            by_date = stats["by_date"]
            if view_mode == "日视图":
                for ds in sorted(by_date.keys()):
                    d = by_date[ds]
                    avg = round(d["final"] / d["count"], 2) if d["count"] > 0 else 0
                    period_rows.append([ds, d["count"], round(d["base"], 2), round(d["discount"], 2),
                                        round(d["final"], 2), round(d["hours"], 1), avg])
            elif view_mode == "周视图":
                from datetime import timedelta as td
                weeks = {}
                for ds in sorted(by_date.keys()):
                    d_date = date.fromisoformat(ds)
                    monday = d_date - td(days=d_date.weekday())
                    sunday = monday + td(days=6)
                    week_key = f"{monday.isoformat()} ~ {sunday.isoformat()}"
                    if week_key not in weeks:
                        weeks[week_key] = {"count": 0, "base": 0.0, "discount": 0.0,
                                            "final": 0.0, "hours": 0.0}
                    for k in ["count", "base", "discount", "final", "hours"]:
                        weeks[week_key][k] += by_date[ds][k]
                for week_key in sorted(weeks.keys()):
                    w = weeks[week_key]
                    avg = round(w["final"] / w["count"], 2) if w["count"] > 0 else 0
                    period_rows.append([week_key, w["count"], round(w["base"], 2), round(w["discount"], 2),
                                        round(w["final"], 2), round(w["hours"], 1), avg])
            elif view_mode == "月视图":
                months = {}
                for ds in sorted(by_date.keys()):
                    month_key = ds[:7]
                    if month_key not in months:
                        months[month_key] = {"count": 0, "base": 0.0, "discount": 0.0,
                                              "final": 0.0, "hours": 0.0}
                    for k in ["count", "base", "discount", "final", "hours"]:
                        months[month_key][k] += by_date[ds][k]
                for month_key in sorted(months.keys()):
                    m = months[month_key]
                    avg = round(m["final"] / m["count"], 2) if m["count"] > 0 else 0
                    period_rows.append([month_key, m["count"], round(m["base"], 2), round(m["discount"], 2),
                                        round(m["final"], 2), round(m["hours"], 1), avg])

            _style_data_rows(ws1, period_rows, start_row=start_row, num_fmt_cols=[2, 3, 4, 6])

            if analysis:
                start_row2 = start_row + len(period_rows) + 2
                ws1.cell(row=start_row2, column=1, value="经营分析").font = section_font
                start_row2 += 1
                kpi_headers = ["指标", "数值"]
                _style_header_row(ws1, kpi_headers, row=start_row2)
                start_row2 += 1
                kpi_data = [
                    ["桌台利用率", f"{analysis['utilization']}%"],
                    ["统计天数", analysis['date_range_days']],
                    ["桌台总数", analysis['total_tables']],
                    ["会员消费占比(笔数)", f"{analysis['member_ratio']}%"],
                    ["会员消费占比(金额)", f"{analysis['member_amount_ratio']}%"],
                    ["累计节省金额(全部会员)",
                     round(sum(d["amount"] for d in analysis["by_discount_source"].values()), 2)],
                ]
                _style_data_rows(ws1, kpi_data, start_row=start_row2, num_fmt_cols=[1])

            _auto_width(ws1, 7)

            # ============ Sheet2: 支付方式 ============
            ws2 = wb.create_sheet("支付方式")
            _write_title(ws2, "茶楼麻将房 - 支付方式统计表", period_str, 5)

            ws2.cell(row=4, column=1, value="按支付方式统计").font = section_font
            pay_headers = ["支付方式", "笔数", "金额", "使用时长(h)", "占比"]
            _style_header_row(ws2, pay_headers, row=5)

            total_final = stats['total_final'] if stats['total_final'] > 0 else 1
            pay_rows = []
            for method, data in sorted(stats["by_payment"].items()):
                pct = round(data["amount"] / total_final * 100, 1)
                pay_rows.append([method, data["count"], round(data["amount"], 2),
                                 round(data["hours"], 1), f"{pct}%"])
            _style_data_rows(ws2, pay_rows, start_row=6, num_fmt_cols=[2])

            if analysis:
                start_row3 = 6 + len(pay_rows) + 2
                ws2.cell(row=start_row3, column=1, value="按优惠来源统计").font = section_font
                start_row3 += 1
                src_headers = ["优惠来源", "使用笔数", "优惠金额"]
                _style_header_row(ws2, src_headers, row=start_row3)
                start_row3 += 1
                src_rows = []
                for src in ["会员折扣", "优惠券折扣"]:
                    d = analysis["by_discount_source"].get(src, {"count": 0, "amount": 0})
                    src_rows.append([src, d["count"], round(d["amount"], 2)])
                _style_data_rows(ws2, src_rows, start_row=start_row3, num_fmt_cols=[2])

            _auto_width(ws2, 5)

            # ============ Sheet3: 会员消费 ============
            ws3 = wb.create_sheet("会员消费")
            _write_title(ws3, "茶楼麻将房 - 会员消费统计", period_str, 7)

            if analysis:
                ws3.cell(row=4, column=1, value="会员消费总体分析").font = section_font
                mem_headers = ["指标", "数值"]
                _style_header_row(ws3, mem_headers, row=5)
                mem_rows = [
                    ["会员账单数", analysis["member_count"]],
                    ["会员实收金额", round(analysis["member_amount"], 2)],
                    ["会员笔数占比", f"{analysis['member_ratio']}%"],
                    ["会员金额占比", f"{analysis['member_amount_ratio']}%"],
                    ["会员折扣节省(全部)",
                     round(analysis["by_discount_source"].get("会员折扣", {"amount": 0})["amount"], 2)],
                    ["优惠券节省(会员单)",
                     round(analysis["by_discount_source"].get("优惠券折扣", {"amount": 0})["amount"], 2)],
                ]
                _style_data_rows(ws3, mem_rows, start_row=6, num_fmt_cols=[1])

                start_row4 = 6 + len(mem_rows) + 2
                ws3.cell(row=start_row4, column=1, value="各账单会员消费明细（全部账单）").font = section_font
                start_row4 += 1
                detail_bill_headers = ["账单号", "桌号", "会员ID", "应收", "优惠", "实收", "支付方式", "支付时间"]
                _style_header_row(ws3, detail_bill_headers, row=start_row4)
                start_row4 += 1

                from database.models import Bill as BillModel
                bills_all = self.db.query(BillModel).filter(
                    BillModel.is_paid == True,
                    BillModel.paid_at >= datetime.combine(analysis["start_date"], datetime.min.time()),
                    BillModel.paid_at <= datetime.combine(analysis["end_date"], datetime.max.time())
                ).order_by(BillModel.paid_at.desc()).all()
                bill_rows = []
                for b in bills_all:
                    bill_rows.append([
                        b.bill_no,
                        b.table_number or "",
                        b.member_id or "非会员",
                        round(b.base_amount, 2),
                        round(b.discount_amount, 2),
                        round(b.final_amount, 2),
                        b.payment_method or "",
                        b.paid_at.strftime('%Y-%m-%d %H:%M:%S') if b.paid_at else ""
                    ])
                _style_data_rows(ws3, bill_rows, start_row=start_row4, num_fmt_cols=[3, 4, 5])

            _auto_width(ws3, 8, min_w=16)

            # ============ Sheet4: 桌台利用率 ============
            ws4 = wb.create_sheet("桌台利用率")
            _write_title(ws4, "茶楼麻将房 - 桌台利用率", period_str, 6)

            if analysis:
                ws4.cell(row=4, column=1, value="桌台使用明细").font = section_font
                util_headers = ["桌号", "消费次数", "营收金额", "使用时长(h)", "估算利用率", "排名"]
                _style_header_row(ws4, util_headers, row=5)

                possible_per_table = analysis["date_range_days"] * 12
                sorted_tables = sorted(analysis["by_table"].items(),
                                        key=lambda x: x[1]["hours"], reverse=True)
                table_rows = []
                for rank, (tn, t) in enumerate(sorted_tables, 1):
                    util_e = round(t["hours"] / possible_per_table * 100, 1) if possible_per_table > 0 else 0
                    table_rows.append([
                        tn, t["count"], round(t["final"], 2),
                        round(t["hours"], 1), f"{util_e}%", rank
                    ])
                _style_data_rows(ws4, table_rows, start_row=6, num_fmt_cols=[2])

                start_row5 = 6 + len(table_rows) + 2
                ws4.cell(row=start_row5, column=1, value="热门时段(按到店次数)").font = section_font
                start_row5 += 1
                hour_headers = ["时段", "次数", "营收金额", "排名"]
                _style_header_row(ws4, hour_headers, row=start_row5)
                start_row5 += 1
                hour_rows = []
                for i, hh in enumerate(analysis["hot_hours"]):
                    hour_rows.append([hh["hour_range"], hh["count"], round(hh["amount"], 2), i + 1])
                _style_data_rows(ws4, hour_rows, start_row=start_row5, num_fmt_cols=[2])

                start_row6 = start_row5 + len(hour_rows) + 2
                ws4.cell(row=start_row6, column=1, value="利用率汇总指标").font = section_font
                start_row6 += 1
                kpi_h = ["指标", "数值"]
                _style_header_row(ws4, kpi_h, row=start_row6)
                start_row6 += 1
                avg_hours = round(analysis["total_hours"] / analysis["total_tables"], 1) if analysis["total_tables"] > 0 else 0
                util_kpi = [
                    ["统计天数", analysis["date_range_days"]],
                    ["桌台总数", analysis["total_tables"]],
                    ["全部桌台总时长(h)", round(analysis["total_hours"], 1)],
                    ["单桌平均使用时长(h)", avg_hours],
                    ["每日桌台平均次数", round(analysis["bill_count"] / analysis["date_range_days"], 1) if analysis["date_range_days"] > 0 else 0],
                    ["整体桌台利用率", f"{analysis['utilization']}%"],
                ]
                _style_data_rows(ws4, util_kpi, start_row=start_row6, num_fmt_cols=[1])

            _auto_width(ws4, 6)

            wb.save(file_path)
            self.show_info("成功", f"报表已导出到:\n{file_path}\n共 4 个工作表:营业汇总/支付方式/会员消费/桌台利用率")
        except ImportError:
            self.show_error("错误", "导出Excel需要安装 openpyxl 库")
        except Exception as e:
            self.show_error("错误", f"导出失败: {str(e)}")
            import traceback
            traceback.print_exc()

    def refresh(self):
        self.refresh_bills()
        self.on_bill_selected()

    def refresh_bills(self):
        self.clear_table(self.bill_widget)
        start_date = self.date_start.date().toPython()
        end_date = self.date_end.date().toPython()
        only_unpaid = self.chk_only_unpaid.isChecked()
        bills = self.bill_service.get_all_bills(start_date, end_date, only_unpaid)

        total_base = 0
        total_discount = 0
        total_final = 0
        total_paid = 0

        for bill in bills:
            row = self.bill_widget.rowCount()
            self.bill_widget.insertRow(row)

            status = "已支付" if bill.is_paid else "未支付"
            payment_method = bill.payment_method or "-"
            created_date = bill.created_at.strftime('%Y-%m-%d %H:%M') if bill.created_at else ""

            self.set_table_row(self.bill_widget, row, [
                bill.id, bill.bill_no, bill.table_number, bill.customer_name,
                created_date, f"¥{bill.base_amount:.2f}", f"¥{bill.discount_amount:.2f}",
                f"¥{bill.final_amount:.2f}", payment_method, status
            ])

            status_item = self.bill_widget.item(row, 9)
            if bill.is_paid:
                status_item.setForeground(QBrush(QColor("#4caf50")))
            else:
                status_item.setForeground(QBrush(QColor("#f44336")))

            total_base += bill.base_amount
            total_discount += bill.discount_amount
            total_final += bill.final_amount
            if bill.is_paid:
                total_paid += bill.final_amount

        self.lbl_stats.setText(
            f"共 {len(bills)} 笔 | 营业额: ¥{total_base:.2f} | 优惠: ¥{total_discount:.2f} | 实收: ¥{total_paid:.2f}"
        )

    def on_bill_selected(self):
        pass

    def get_selected_bill_id(self):
        row = self.bill_widget.currentRow()
        if row >= 0:
            return int(self.bill_widget.item(row, 0).text())
        return None

    def view_bill(self):
        bill_id = self.get_selected_bill_id()
        if not bill_id:
            self.show_info("提示", "请选择要查看的账单")
            return
        content = self.bill_service.generate_print_content(bill_id)
        dialog = QDialog(self)
        dialog.setWindowTitle("账单详情")
        dialog.setModal(True)
        dialog.resize(450, 600)
        layout = QVBoxLayout(dialog)
        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setPlainText(content)
        text_edit.setStyleSheet("font-family: Consolas, Monaco, monospace;")
        layout.addWidget(text_edit)
        dialog.exec()

    def print_bill(self):
        bill_id = self.get_selected_bill_id()
        if not bill_id:
            self.show_info("提示", "请选择要打印的账单")
            return

        try:
            content = self.bill_service.generate_print_content(bill_id)
            from PySide6.QtPrintSupport import QPrinter, QPrintDialog
            from PySide6.QtGui import QTextDocument

            printer = QPrinter(QPrinter.HighResolution)
            printer.setPageSize(QPrinter.A4)
            printer.setOutputFormat(QPrinter.NativeFormat)
            dialog = QPrintDialog(printer, self)
            if dialog.exec() == QDialog.Accepted:
                doc = QTextDocument()
                doc.setPlainText(content)
                doc.print_(printer)
                self.show_info("成功", "账单已发送打印")
        except ImportError:
            self.show_info("提示", "打印功能需要安装 PySide6 的打印支持模块")
        except Exception as e:
            self.show_error("错误", str(e))

    def pay_bill(self):
        bill_id = self.get_selected_bill_id()
        if not bill_id:
            self.show_info("提示", "请选择要支付的账单")
            return

        bill = self.bill_service.get_bill(bill_id)
        if bill and bill.is_paid:
            self.show_info("提示", "该账单已支付")
            return

        payment_methods = ["现金", "微信支付", "支付宝", "银行卡", "会员卡", "其他"]
        method, ok = QInputDialog.getItem(self, "选择支付方式", "请选择支付方式:", payment_methods, 0, False)
        if ok:
            try:
                member_id = None
                if method == "会员卡":
                    member_service = MemberService(self.db)
                    members = member_service.get_all_members(only_active=True)
                    if not members:
                        self.show_error("错误", "没有可用会员，请先在会员档案中添加会员")
                        return
                    member_names = [f"{m.name} ({m.phone}) 余额:¥{m.balance:.2f}" for m in members]
                    member_choice, mok = QInputDialog.getItem(
                        self, "选择会员", "请选择支付会员:", member_names, 0, False)
                    if not mok:
                        return
                    idx = member_names.index(member_choice)
                    member = members[idx]
                    if member.balance < bill.final_amount:
                        self.show_error("余额不足",
                            f"会员「{member.name}」余额 ¥{member.balance:.2f}，"
                            f"应付 ¥{bill.final_amount:.2f}，差额 ¥{bill.final_amount - member.balance:.2f}")
                        return
                    member_id = member.id

                self.bill_service.pay_bill(bill_id, method, member_id=member_id)
                self.show_info("成功", "支付成功")
                self.refresh()
            except Exception as e:
                self.show_error("错误", str(e))

    def delete_bill(self):
        bill_id = self.get_selected_bill_id()
        if not bill_id:
            self.show_info("提示", "请选择要删除的账单")
            return
        if self.show_confirm("确认", "确定要删除这个账单吗？"):
            if self.bill_service.delete_bill(bill_id):
                self.show_info("成功", "账单删除成功")
                self.refresh()
            else:
                self.show_error("错误", "删除失败")
